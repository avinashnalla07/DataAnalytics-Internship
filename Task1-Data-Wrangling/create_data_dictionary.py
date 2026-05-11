import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "data_dictionary.xlsx")

# All 27 columns with their details
columns_data = [
    ("Row ID", "int64", "Unique row number for each record", "Helps in referencing specific rows in the dataset"),
    ("Order ID", "object (string)", "Unique ID for each order placed", "Tracks individual orders; one order can have multiple items"),
    ("Order Date", "datetime64", "Date when the customer placed the order", "Needed for trend analysis and finding seasonal patterns"),
    ("Ship Date", "datetime64", "Date when the order was shipped out", "Helps measure how fast orders get fulfilled"),
    ("Ship Mode", "object (string)", "Shipping method: Standard, Second, First, or Same Day", "Affects delivery cost and customer satisfaction"),
    ("Customer ID", "object (string)", "Unique ID for each customer", "Connects orders to customers for segmentation"),
    ("Customer Name", "object (string)", "Full name of the customer", "Used for customer profiling"),
    ("Segment", "object (string)", "Customer type: Consumer, Corporate, or Home Office", "Useful for understanding which customer group buys what"),
    ("Country", "object (string)", "Country of the order (United States only)", "All orders are from the US in this dataset"),
    ("City", "object (string)", "City where the order was delivered", "Helps with city-level sales analysis"),
    ("State", "object (string)", "State where the order was delivered", "Important for regional comparisons and tax analysis"),
    ("Postal Code", "int64", "ZIP code of the delivery address", "Can be used for location-based analysis"),
    ("Region", "object (string)", "Geographic region: East, West, Central, or South", "Used to compare performance across different parts of the country"),
    ("Product ID", "object (string)", "Unique ID for each product", "Links products across multiple orders"),
    ("Category", "object (string)", "Product category: Furniture, Office Supplies, or Technology", "Main way to group products for analysis"),
    ("Sub-Category", "object (string)", "More specific product type (e.g., Chairs, Phones, Binders)", "Gives a more detailed view of product performance"),
    ("Product Name", "object (string)", "Full name of the product", "Identifies the exact product sold"),
    ("Sales", "float64", "Revenue from the sale in USD", "The main revenue number; drives most of the financial analysis"),
    ("Quantity", "int64", "Number of units sold", "Shows volume; useful for inventory planning"),
    ("Discount", "float64", "Discount applied (0.0 to 0.8, i.e., 0% to 80%)", "Directly impacts profit; key for pricing decisions"),
    ("Profit", "float64", "Net profit in USD (can be negative)", "The bottom line; tells us if we actually made money"),
    ("Days_to_Ship", "int64", "Days between Order Date and Ship Date", "Measures how quickly orders get shipped out"),
    ("Profit_Margin_%", "float64", "Profit as a percentage of Sales", "Makes it easy to compare profitability across different products"),
    ("Order_Year", "int64", "Year extracted from Order Date (2014-2017)", "Used for year-over-year comparisons"),
    ("Order_Month", "int64", "Month number from Order Date (1-12)", "Helps spot monthly trends and seasonal patterns"),
    ("Order_Month_Name", "object (string)", "Month name (January, February, etc.)", "More readable version of the month for reports"),
    ("Is_Loss", "bool", "True if Profit is negative, False otherwise", "Quick way to filter and count loss-making orders"),
]

# Create the Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Data Dictionary"

# Styling
header_font = Font(name='Calibri', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='0A1628', end_color='0A1628', fill_type='solid')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
cell_font = Font(name='Calibri', size=11)
cell_alignment = Alignment(vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Title row
ws.merge_cells('A1:D1')
title_cell = ws['A1']
title_cell.value = "Superstore Sales Dataset - Data Dictionary"
title_cell.font = Font(name='Calibri', bold=True, size=16, color='0D9488')
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 40

# Subtitle row
ws.merge_cells('A2:D2')
subtitle_cell = ws['A2']
subtitle_cell.value = "Intern: Avinash Nalla | ID: APSPL2632585 | ApexPlanet Software Pvt. Ltd."
subtitle_cell.font = Font(name='Calibri', size=11, italic=True, color='666666')
subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[2].height = 25

# Empty row for spacing
ws.row_dimensions[3].height = 10

# Column headers
headers = ['Column Name', 'Data Type', 'Description', 'Business Relevance']
for col_idx, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border
ws.row_dimensions[4].height = 30

# Data rows with alternating colors
alt_fill = PatternFill(start_color='F0FDFA', end_color='F0FDFA', fill_type='solid')
for row_idx, (col_name, dtype, desc, relevance) in enumerate(columns_data, 5):
    values = [col_name, dtype, desc, relevance]
    for col_idx, value in enumerate(values, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = cell_font
        cell.alignment = cell_alignment
        cell.border = thin_border
        if row_idx % 2 == 1:
            cell.fill = alt_fill
    ws.row_dimensions[row_idx].height = 45

# Set column widths
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 55
ws.column_dimensions['D'].width = 55

# Summary row at the bottom
summary_row = 5 + len(columns_data) + 1
ws.merge_cells(f'A{summary_row}:D{summary_row}')
summary_cell = ws.cell(row=summary_row, column=1, 
    value=f"Total Columns: {len(columns_data)} | Dataset: 9,994 rows | Source: Superstore Sales Dataset")
summary_cell.font = Font(name='Calibri', bold=True, size=11, color='0D9488')
summary_cell.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[summary_row].height = 30

wb.save(OUTPUT_FILE)
print(f"data_dictionary.xlsx created at: {OUTPUT_FILE}")
print(f"Total columns documented: {len(columns_data)}")
